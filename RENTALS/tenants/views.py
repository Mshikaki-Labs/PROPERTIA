from datetime import date
import csv
import io

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.contrib.auth.decorators import login_required

from .models import Tenant
from .forms import TenantForm
from invoices.models import Invoice
from payments.models import Payment
from properties.models import Property
from units.models import Unit
from PROPATIA.pagination import paginate_queryset
from accounts.access_utils import get_accessible_properties, filter_tenants_by_accessible_properties


@login_required
def tenant_list(request):
    if request.method == "POST":
        form = TenantForm(request.POST, user=request.user)
        if form.is_valid():
            form.save()
            return redirect('tenants:tenant_list')
    else:
        form = TenantForm(user=request.user)

    accessible_props = get_accessible_properties(request.user)
    tenants = Tenant.objects.filter(
        unit__property__in=accessible_props
    ).select_related('unit', 'unit__property').order_by('last_name', 'first_name')

    property_filter = request.GET.get('property')
    status_filter = request.GET.get('status')
    allowed_property_ids = set(accessible_props.values_list('id', flat=True))

    if property_filter and property_filter.isdigit() and int(property_filter) in allowed_property_ids:
        tenants = tenants.filter(unit__property_id=property_filter)
    else:
        property_filter = None

    if status_filter:
        tenants = tenants.filter(status=status_filter)

    pagination = paginate_queryset(request, tenants)

    return render(request, 'tenants/tenants_view.html', {
        'tenants': pagination['page_obj'],
        'form': form,
        'properties': accessible_props,
        'selected_property': property_filter,
        'selected_status': status_filter,
        **pagination,
    })


@login_required
def available_units(request, pk):
    """Return units for a property as JSON."""
    accessible_props = get_accessible_properties(request.user)
    property_obj = get_object_or_404(accessible_props, pk=pk)
    units = Unit.objects.filter(property=property_obj).values('id', 'name', 'rent_amount')
    return JsonResponse({'property': property_obj.name, 'units': list(units)})


@login_required
def tenant_ledger(request, pk):
    """Render the tenant ledger with invoices and payments for the tenant/unit."""
    accessible_props = get_accessible_properties(request.user)
    tenant = get_object_or_404(
        Tenant.objects.select_related('unit', 'unit__property'),
        pk=pk,
        unit__property__in=accessible_props,
    )

    invoices = Invoice.objects.filter(tenant=tenant).select_related('unit').order_by('due_date', 'id')
    payments = Payment.objects.filter(tenant=tenant).select_related('unit').order_by('date', 'id')
    tenant_property = tenant.unit.property if tenant.unit else None
    property_tenants = Tenant.objects.filter(
        unit__property=tenant_property,
    ).select_related('unit').order_by('unit__name', 'last_name', 'first_name')

    ledger_entries = []
    for invoice in invoices:
        ledger_entries.append({
            'date': invoice.due_date,
            'code': '',
            'description': f"Invoice {invoice.invoice_number}",
            'debit': invoice.amount,
            'credit': 0,
            'status': invoice.status,
            'is_overdue': invoice.due_date < date.today() and invoice.status != 'Paid',
            'type': 'invoice',
        })

    for payment in payments:
        ledger_entries.append({
            'date': payment.date,
            'code': payment.code or '',
            'description': payment.description or 'Payment received',
            'debit': 0,
            'credit': payment.amount,
            'status': payment.status,
            'is_overdue': False,
            'type': 'payment',
        })

    ledger_entries.sort(key=lambda item: (item['date'], item['type'] != 'invoice'))

    running_balance = 0
    for entry in ledger_entries:
        running_balance += entry['debit'] - entry['credit']
        entry['balance'] = running_balance

    total_invoiced = sum(invoice.amount for invoice in invoices)
    total_paid = sum(payment.amount for payment in payments)
    current_balance = total_invoiced - total_paid
    overdue_count = sum(
        1 for invoice in invoices if invoice.due_date < date.today() and invoice.status != 'Paid'
    )

    return render(request, 'tenants/tenant_ledger.html', {
        'tenant': tenant,
        'ledger_entries': ledger_entries,
        'total_invoiced': total_invoiced,
        'total_paid': total_paid,
        'current_balance': current_balance,
        'overdue_count': overdue_count,
        'property_tenants': property_tenants,
    })


@login_required
def delete_tenants(request):
    """Delete selected tenants."""
    if request.method == 'POST':
        accessible_props = get_accessible_properties(request.user)
        tenant_ids = request.POST.getlist('tenant_ids[]')
        deleted_count, _ = Tenant.objects.filter(
            id__in=tenant_ids,
            unit__property__in=accessible_props,
        ).delete()
        return JsonResponse({'success': True, 'message': f'{deleted_count} tenant(s) deleted'})
    return JsonResponse({'success': False, 'message': 'Invalid request'})


@login_required
def edit_tenant(request, pk):
    """Edit an existing tenant."""
    accessible_props = get_accessible_properties(request.user)
    tenant = get_object_or_404(
        Tenant.objects.select_related('unit', 'unit__property'),
        pk=pk,
        unit__property__in=accessible_props,
    )
    if request.method == 'POST':
        form = TenantForm(request.POST, request.FILES, instance=tenant, user=request.user)
        if form.is_valid():
            form.save()
            return redirect('tenants:tenant_list')
    else:
        form = TenantForm(instance=tenant, user=request.user)

    return render(request, 'tenants/edit_tenant.html', {'form': form, 'tenant': tenant})


@login_required
def upload_tenants(request):
    """Upload tenants from CSV or XLSX file"""
    def normalize_header(header):
        return ' '.join(
            str(header)
            .replace('\ufeff', '')
            .replace('*', '')
            .replace('\\', '')
            .lower()
            .replace('_', ' ')
            .split()
        )

    def get_row_value(row, *headers):
        for header in headers:
            value = row.get(normalize_header(header))
            if value is not None and str(value).strip() != '':
                return value
        return None

    def normalize_upload_row(headers, values):
        row = {}
        for idx, header in enumerate(headers):
            normalized_header = normalize_header(header)
            if not normalized_header:
                continue
            value = values[idx] if idx < len(values) else None
            current_value = row.get(normalized_header)
            if current_value is None or str(current_value).strip() == '':
                row[normalized_header] = value
        return row

    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file provided'})

    file = request.FILES['file']
    filename = file.name.lower()
    validate_only = request.POST.get('validate_only') == '1'

    try:
        rows = []

        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            sample = content[:2048]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t;')
            except csv.Error:
                dialect = csv.excel_tab if '\t' in sample else csv.excel
            reader = csv.reader(content.splitlines(), dialect=dialect)
            headers = next(reader, [])
            for line in reader:
                if any(str(v).strip() for v in line):
                    rows.append(normalize_upload_row(headers, line))
        elif filename.endswith('.xlsx'):
            try:
                from openpyxl import load_workbook
            except ImportError:
                return JsonResponse({'success': False, 'message': 'openpyxl is not installed. Please use CSV format or contact admin.'})
            wb = load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None and str(v).strip() != '' for v in row):
                    rows.append(normalize_upload_row(headers, list(row)))
        else:
            return JsonResponse({'success': False, 'message': 'Invalid file format. Please use CSV or XLSX'})

        if not rows:
            return JsonResponse({'success': False, 'message': 'File is empty'})

        accessible_props = get_accessible_properties(request.user)
        created_count = 0
        errors = []

        for idx, row in enumerate(rows, start=2):
            try:
                first_name = get_row_value(row, 'first name', 'firstname', 'first')
                last_name = get_row_value(row, 'last name', 'lastname', 'last')
                phone_number = get_row_value(row, 'phone number', 'phone', 'phone_number')
                property_name = get_row_value(row, 'property')
                next_of_kin_name = get_row_value(row, 'next of kin name', 'next_of_kin', 'kin name')
                next_of_kin_phone = get_row_value(row, 'next of kin phone', 'next_of_kin_phone_number', 'kin phone')
                description = get_row_value(row, 'description')

                if not first_name or not last_name or not phone_number or not property_name:
                    errors.append(f'Row {idx}: Missing required fields (first_name, last_name, phone_number, property)')
                    continue

                property_obj = accessible_props.filter(name__iexact=str(property_name).strip()).first()
                if not property_obj:
                    errors.append(f'Row {idx}: Property "{property_name}" not found or not accessible')
                    continue

                unit = Unit.objects.filter(property=property_obj, status='vacant').first()
                if not unit:
                    errors.append(f'Row {idx}: No vacant unit available in property "{property_name}"')
                    continue

                tenant_data = {
                    'unit': unit,
                    'first_name': str(first_name).strip(),
                    'last_name': str(last_name).strip(),
                    'phone_number': str(phone_number).strip(),
                    'next_of_kin_name': str(next_of_kin_name).strip() if next_of_kin_name else '',
                    'next_of_kin_phone_number': str(next_of_kin_phone).strip() if next_of_kin_phone else '',
                    'description': str(description).strip() if description else '',
                    'status': 'active',
                    'deposit_required': False,
                    'deposit_amount': 0,
                }

                if validate_only:
                    created_count += 1
                else:
                    Tenant.objects.create(**tenant_data)
                    created_count += 1

            except ValueError as ve:
                errors.append(f'Row {idx}: Invalid data format - {str(ve)}')
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')

        message = f'Successfully processed {created_count} tenants'
        if errors:
            message += f'. {len(errors)} errors: ' + '; '.join(errors[:3])

        response = {
            'success': created_count > 0 or not errors,
            'count': created_count,
            'message': message,
            'errors': errors,
            'invalid_rows': len(errors),
            'valid_rows': created_count,
        }
        return JsonResponse(response)

    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error processing file: {str(e)}'})
