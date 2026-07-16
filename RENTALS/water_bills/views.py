import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.views.decorators.http import require_POST


from .models import WaterBill, WaterBillPayment, WaterBillPaymentAllocation
from properties.models import Property
from units.models import Unit
from PROPATIA.pagination import paginate_queryset


def normalize_header(header):
    return ' '.join(
        str(header)
        .replace('\ufeff', '')
        .replace('*', '')
        .replace('\\', '')
        .lower()
        .replace('_', ' ')
        .split()
    )


def get_row_value(row, *headers):
    for header in headers:
        value = row.get(normalize_header(header))
        if value is not None and str(value).strip() != '':
            return value
    return None


def normalize_upload_row(headers, values):
    row = {}
    for idx, header in enumerate(headers):
        normalized_header = normalize_header(header)
        if not normalized_header:
            continue

        value = values[idx] if idx < len(values) else None
        current_value = row.get(normalized_header)
        if current_value is None or str(current_value).strip() == '':
            row[normalized_header] = value
    return row


def parse_decimal(value):
    if value is None or str(value).strip() == '':
        raise ValueError('Missing amount')
    value_text = str(value).strip().replace(',', '')
    value_text = re.sub(r'[^\d.\-()]', '', value_text)
    if value_text.startswith('(') and value_text.endswith(')'):
        value_text = f'-{value_text[1:-1]}'
    if not value_text:
        raise ValueError('Invalid amount')
    try:
        return Decimal(value_text)
    except (InvalidOperation, ValueError):
        raise ValueError('Invalid amount')


def parse_integer(value, field_name):
    if value is None or str(value).strip() == '':
        raise ValueError(f'Missing {field_name}')
    value_text = str(value).strip().replace(',', '')
    try:
        decimal_value = Decimal(value_text)
    except (InvalidOperation, ValueError):
        raise ValueError(f'Invalid {field_name}')
    if decimal_value != decimal_value.to_integral_value():
        raise ValueError(f'{field_name} must be a whole number')
    return int(decimal_value)


def parse_payment_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%m-%d-%Y', '%d.%m.%Y']:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    raise ValueError('Invalid date format')


def parse_optional_water_bill_due_date(value):
    if value is None or str(value).strip() == '':
        return date.today()
    return parse_payment_date(value)


def normalize_water_bill_status(value):
    if value is None or str(value).strip() == '':
        return 'Unpaid'
    status = str(value).strip().lower()
    if status == 'paid':
        return 'Paid'
    if status == 'unpaid':
        return 'Unpaid'
    raise ValueError('Invalid status')


def read_upload_rows(uploaded_file):
    filename = uploaded_file.name.lower()
    if filename.endswith('.csv'):
        content = uploaded_file.read().decode('utf-8-sig')
        sample = content[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',\t;')
        except csv.Error:
            dialect = csv.excel_tab if '\t' in sample else csv.excel
        reader = csv.reader(content.splitlines(), dialect=dialect)
        headers = next(reader, [])
        return [
            normalize_upload_row(headers, row)
            for row in reader
            if any(value is not None and str(value).strip() != '' for value in row)
        ]

    if filename.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError('openpyxl is not installed. Please use CSV format or contact admin.')

        wb = load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = normalize_upload_row(headers, row)
            if any(value is not None and str(value).strip() != '' for value in row_dict.values()):
                rows.append(row_dict)
        return rows

    raise ValueError('Invalid file format. Please use CSV or XLSX')


def resolve_water_bill_tenant(unit):
    active_lease = unit.leases.filter(is_active=True).select_related('tenant').first()
    if active_lease and active_lease.tenant:
        return active_lease.tenant
    return unit.tenants.filter(status='active').first() or unit.tenants.first()


def set_water_payment_status_from_balance(payment):
    payment.status = 'claimed' if Decimal(payment.balance or 0) <= 0 else 'unclaimed'
    payment.save(update_fields=['balance', 'status'])


def is_app_admin(user):
    return user.is_authenticated and hasattr(user, 'profile') and user.profile.role == 'admin'


@transaction.atomic
def allocate_payment_to_water_bills(payment):
    payment = WaterBillPayment.objects.select_for_update().get(pk=payment.pk)
    if Decimal(payment.balance or 0) <= 0:
        return []

    allocations = []
    water_bills = WaterBill.objects.select_for_update().filter(
        user=payment.user,
        unit=payment.unit,
        tenant=payment.tenant,
    ).exclude(status='Paid').order_by('due_date', 'id')

    for water_bill in water_bills:
        remaining = Decimal(water_bill.get_remaining_balance() or 0)
        if remaining <= 0:
            water_bill.update_status()
            continue

        amount = min(Decimal(payment.balance or 0), remaining)
        if amount <= 0:
            break

        allocation, created = WaterBillPaymentAllocation.objects.get_or_create(
            water_bill=water_bill,
            payment=payment,
            defaults={'amount_applied': amount},
        )
        if not created:
            continue

        allocations.append(allocation)
        payment.balance = Decimal(payment.balance or 0) - amount
        set_water_payment_status_from_balance(payment)
        water_bill.update_status()

        if Decimal(payment.balance or 0) <= 0:
            break

    return allocations

@login_required
def water_bill_list(request):
    if request.method == "POST":
        unit_id = request.POST.get('unit')
        prev = int(request.POST.get('previous_reading'))
        curr = int(request.POST.get('current_reading'))
        rate = float(request.POST.get('rate'))
        date = request.POST.get('due_date')
        
        unit = Unit.objects.get(id=unit_id, user=request.user)
        tenant = unit.tenants.filter(status='active').first()
        
        if tenant:
            WaterBill.objects.create(
                user=request.user,
                unit=unit,
                tenant=tenant,
                previous_reading=prev,
                current_reading=curr,
                rate=rate,
                due_date=date
            )
        return redirect('water_bills:water_bills_home')

    selected_property = request.GET.get('property', '')
    selected_unit = request.GET.get('unit', '')
    selected_status = request.GET.get('status', '')
    selected_start_date = request.GET.get('start_date', '')
    selected_end_date = request.GET.get('end_date', '')

    bills = WaterBill.objects.filter(user=request.user).select_related('unit', 'unit__property', 'tenant').order_by('-due_date')

    if selected_property:
        bills = bills.filter(unit__property_id=selected_property)
    if selected_unit:
        bills = bills.filter(unit_id=selected_unit)
    if selected_status:
        bills = bills.filter(status=selected_status)
    if selected_start_date:
        bills = bills.filter(due_date__gte=selected_start_date)
    if selected_end_date:
        bills = bills.filter(due_date__lte=selected_end_date)

    pagination = paginate_queryset(request, bills)

    context = {
        'bills': pagination['page_obj'],
        'properties': Property.objects.filter(user=request.user),
        'units': Unit.objects.filter(user=request.user).select_related('property').order_by('property__name', 'name'),
        'occupied_units': Unit.objects.filter(user=request.user, status='occupied').select_related('property').order_by('property__name', 'name'),
        'can_delete_water_bills': is_app_admin(request.user),
        'selected_property': selected_property,
        'selected_unit': selected_unit,
        'selected_status': selected_status,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
    }
    context.update(pagination)
    return render(request, 'water_bills/water_bills_view.html', context)


@login_required
@require_POST
@transaction.atomic
def delete_water_bills(request):
    if not is_app_admin(request.user):
        return JsonResponse({'success': False, 'message': 'Only admins can delete water bills'}, status=403)

    bill_ids = request.POST.getlist('bill_ids[]')
    if not bill_ids:
        return JsonResponse({'success': False, 'message': 'No water bills selected'})

    bills = WaterBill.objects.filter(id__in=bill_ids, user=request.user)
    bill_count = bills.count()
    allocations = WaterBillPaymentAllocation.objects.filter(water_bill__in=bills).select_related('payment')

    payments_to_restore = {}
    for allocation in allocations:
        payment = allocation.payment
        payment.balance = Decimal(payment.balance or 0) + allocation.amount_applied
        payments_to_restore[payment.id] = payment

    for payment in payments_to_restore.values():
        set_water_payment_status_from_balance(payment)

    bills.delete()

    return JsonResponse({
        'success': True,
        'message': f'Successfully deleted {bill_count} water bill(s)'
    })


@login_required
@require_POST
def bulk_generate_water_bills(request):
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file provided'})

    validate_only = request.POST.get('validate_only') == '1'

    try:
        rows = read_upload_rows(request.FILES['file'])
    except RuntimeError as e:
        return JsonResponse({'success': False, 'message': str(e)})
    except ValueError as e:
        return JsonResponse({'success': False, 'message': str(e)})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error processing file: {str(e)}'})

    if not rows:
        return JsonResponse({'success': False, 'message': 'File is empty'})

    created_count = 0
    errors = []
    valid_rows = 0

    for idx, row in enumerate(rows, start=2):
        try:
            row_lower = {normalize_header(k): v for k, v in row.items()}

            property_name = get_row_value(row_lower, 'property')
            house_number = get_row_value(
                row_lower,
                'house number',
                'HOUSE_NUMBER',
                'house no',
                'house',
                'unit',
                'unit number',
            )
            previous_reading = get_row_value(row_lower, 'previous reading', 'previous_reading')
            current_reading = get_row_value(row_lower, 'current reading', 'current_reading')
            consumption = get_row_value(row_lower, 'consumption')
            rate = get_row_value(row_lower, 'rate', 'rate per unit', 'RATE PER UNIT')
            amount = get_row_value(row_lower, 'amount')
            due_date = get_row_value(row_lower, 'due date', 'date')
            status = get_row_value(row_lower, 'status')

            if not property_name or not house_number or previous_reading is None or current_reading is None or consumption is None or rate is None or amount is None:
                errors.append(f'Row {idx}: Missing required fields (PROPERTY, HOUSE NUMBER, PREVIOUS READING, CURRENT READING, CONSUMPTION, RATE, AMOUNT)')
                continue

            property_obj = Property.objects.filter(name__iexact=property_name, user=request.user).first()
            if not property_obj:
                errors.append(f'Row {idx}: Property "{property_name}" not found for this user')
                continue

            unit_obj = Unit.objects.filter(
                property=property_obj,
                name__iexact=str(house_number).strip(),
                user=request.user,
            ).first()
            if not unit_obj:
                errors.append(f'Row {idx}: House number "{house_number}" not found in property "{property_name}"')
                continue

            tenant_obj = resolve_water_bill_tenant(unit_obj)
            if not tenant_obj:
                errors.append(f'Row {idx}: No active tenant found for house/unit "{unit_obj.name}" in property "{property_obj.name}"')
                continue

            try:
                previous_reading_val = parse_integer(previous_reading, 'previous reading')
                current_reading_val = parse_integer(current_reading, 'current reading')
                consumption_val = parse_integer(consumption, 'consumption')
            except ValueError as e:
                errors.append(f'Row {idx}: {str(e)}')
                continue

            if current_reading_val < previous_reading_val:
                errors.append(f'Row {idx}: Current reading cannot be less than previous reading')
                continue

            calculated_consumption = current_reading_val - previous_reading_val
            if consumption_val != calculated_consumption:
                errors.append(f'Row {idx}: Consumption must equal current reading minus previous reading ({calculated_consumption})')
                continue

            try:
                rate_val = parse_decimal(rate)
                amount_val = parse_decimal(amount)
            except ValueError as e:
                errors.append(f'Row {idx}: {str(e)}')
                continue

            calculated_amount = Decimal(consumption_val) * rate_val
            if amount_val != calculated_amount:
                errors.append(f'Row {idx}: Amount must equal consumption multiplied by rate ({calculated_amount})')
                continue

            try:
                due_date_val = parse_optional_water_bill_due_date(due_date)
            except Exception as e:
                errors.append(f'Row {idx}: Error parsing due date - {str(e)}')
                continue

            try:
                status_val = normalize_water_bill_status(status)
            except ValueError as e:
                errors.append(f'Row {idx}: {str(e)}')
                continue

            duplicate_exists = WaterBill.objects.filter(
                user=request.user,
                unit=unit_obj,
                tenant=tenant_obj,
                previous_reading=previous_reading_val,
                current_reading=current_reading_val,
                rate=rate_val,
                due_date=due_date_val,
            ).exists()
            if duplicate_exists:
                errors.append(f'Row {idx}: Duplicate water bill skipped for {property_obj.name} house {unit_obj.name} on {due_date_val}')
                continue

            valid_rows += 1

            if validate_only:
                continue

            WaterBill.objects.create(
                user=request.user,
                unit=unit_obj,
                tenant=tenant_obj,
                previous_reading=previous_reading_val,
                current_reading=current_reading_val,
                rate=rate_val,
                due_date=due_date_val,
                status=status_val,
            )
            created_count += 1
        except ValueError as e:
            errors.append(f'Row {idx}: Invalid data format - {str(e)}')
        except Exception as e:
            errors.append(f'Row {idx}: {str(e)}')

    if validate_only:
        return JsonResponse({
            'success': len(errors) == 0,
            'valid_rows': valid_rows,
            'invalid_rows': len(errors),
            'errors': errors,
            'message': f'Validation complete: {valid_rows} valid, {len(errors)} invalid.',
        })

    message = f'Successfully generated {created_count} water bill(s)'
    if errors:
        message += f'. {len(errors)} errors: ' + '; '.join(errors[:3])

    return JsonResponse({
        'success': created_count > 0 or not errors,
        'count': created_count,
        'message': message,
        'errors': errors,
        'invalid_rows': len(errors),
    })


@login_required
def water_bill_payment_list(request):
    selected_property = request.GET.get('property', '')
    selected_status = request.GET.get('status', '')
    selected_start_date = request.GET.get('start_date', '')
    selected_end_date = request.GET.get('end_date', '')

    payments = WaterBillPayment.objects.filter(user=request.user).select_related(
        'property',
        'unit',
        'tenant',
    ).order_by('date', 'id')

    if selected_property:
        payments = payments.filter(property_id=selected_property)

    if selected_status:
        payments = payments.filter(status=selected_status)

    if selected_start_date:
        payments = payments.filter(date__gte=selected_start_date)

    if selected_end_date:
        payments = payments.filter(date__lte=selected_end_date)

    pagination = paginate_queryset(request, payments)
    page_payments = list(pagination['page_obj'])
    payment_ids = [payment.id for payment in page_payments]
    allocations_by_payment = {payment_id: [] for payment_id in payment_ids}

    if payment_ids:
        allocations = WaterBillPaymentAllocation.objects.filter(
            payment_id__in=payment_ids,
        ).select_related('water_bill').order_by('water_bill__due_date', 'water_bill__id')
        for allocation in allocations:
            allocations_by_payment.setdefault(allocation.payment_id, []).append(allocation)

    for payment in page_payments:
        allocations = allocations_by_payment.get(payment.id, [])
        payment.allocated_amount = sum(allocation.amount_applied for allocation in allocations)
        payment.remaining_credit = payment.balance
        payment.allocation_summary = ', '.join(
            f'{allocation.water_bill.due_date:%b %Y}: KES {allocation.amount_applied}'
            for allocation in allocations
        )
        payment.allocation_count = len(allocations)

    context = {
        'payments': pagination['page_obj'],
        'properties': Property.objects.filter(user=request.user),
        'can_delete_water_bill_payments': is_app_admin(request.user),
        'selected_property': selected_property,
        'selected_status': selected_status,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
    }
    context.update(pagination)
    return render(request, 'water_bills/water_bill_payments_view.html', context)


@login_required
@require_POST
@transaction.atomic
def delete_water_bill_payments(request):
    if not is_app_admin(request.user):
        return JsonResponse({'success': False, 'message': 'Only admins can delete water bill payments'}, status=403)

    payment_ids = request.POST.getlist('payment_ids[]')
    if not payment_ids:
        return JsonResponse({'success': False, 'message': 'No water bill payments selected'})

    payments = WaterBillPayment.objects.filter(id__in=payment_ids, user=request.user)
    payment_count = payments.count()
    water_bill_ids = list(
        WaterBillPaymentAllocation.objects.filter(payment__in=payments)
        .values_list('water_bill_id', flat=True)
        .distinct()
    )

    payments.delete()

    for water_bill in WaterBill.objects.filter(id__in=water_bill_ids, user=request.user):
        water_bill.update_status()

    return JsonResponse({
        'success': True,
        'message': f'Successfully deleted {payment_count} water bill payment(s)'
    })


@login_required
@require_POST
def upload_water_bill_payments(request):
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file provided'})

    validate_only = request.POST.get('validate_only') == '1'

    try:
        rows = read_upload_rows(request.FILES['file'])
    except RuntimeError as e:
        return JsonResponse({'success': False, 'message': str(e)})
    except ValueError as e:
        return JsonResponse({'success': False, 'message': str(e)})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error processing file: {str(e)}'})

    if not rows:
        return JsonResponse({'success': False, 'message': 'File is empty'})

    created_count = 0
    errors = []
    valid_rows = 0

    for idx, row in enumerate(rows, start=2):
        try:
            row_lower = {normalize_header(k): v for k, v in row.items()}

            property_name = get_row_value(row_lower, 'property')
            house_number = get_row_value(
                row_lower,
                'HOUSE_NUMBER',
                'house numner',
                'house no',
                'house',
                'unit',
                'unit number',
            )
            paid_in_date = get_row_value(row_lower, 'paid in date', 'date', 'payment date')
            code = get_row_value(row_lower, 'code', 'receipt', 'reference') or ''
            details = get_row_value(row_lower, 'details', 'description') or ''
            amount = get_row_value(row_lower, 'amount')

            if not property_name or not house_number or amount is None or not paid_in_date:
                errors.append(f'Row {idx}: Missing required fields (PROPERTY, HOUSE_NUMBER, DATE, AMOUNT)')
                continue

            property_obj = Property.objects.filter(name__iexact=property_name, user=request.user).first()
            if not property_obj:
                errors.append(f'Row {idx}: Property "{property_name}" not found for this user')
                continue

            unit_obj = Unit.objects.filter(
                property=property_obj,
                name__iexact=str(house_number).strip(),
                user=request.user,
            ).first()
            if not unit_obj:
                errors.append(f'Row {idx}: House number "{house_number}" not found in property "{property_name}"')
                continue

            tenant_obj = resolve_water_bill_tenant(unit_obj)
            if not tenant_obj:
                errors.append(f'Row {idx}: No active tenant found for house/unit "{unit_obj.name}" in property "{property_obj.name}"')
                continue

            try:
                amount_val = parse_decimal(amount)
            except ValueError:
                errors.append(f'Row {idx}: Invalid amount format "{amount}"')
                continue

            try:
                date_obj = parse_payment_date(paid_in_date)
            except Exception as e:
                errors.append(f'Row {idx}: Error parsing date - {str(e)}')
                continue

            code = str(code).strip()
            description = str(details).strip()

            duplicate_exists = WaterBillPayment.objects.filter(
                user=request.user,
                property=property_obj,
                unit=unit_obj,
                tenant=tenant_obj,
                amount=amount_val,
                date=date_obj,
                code=code,
                description=description,
            ).exists()
            if duplicate_exists:
                errors.append(f'Row {idx}: Duplicate water bill payment skipped for {property_obj.name} house {unit_obj.name} on {date_obj}')
                continue

            valid_rows += 1

            if validate_only:
                continue

            payment = WaterBillPayment.objects.create(
                user=request.user,
                property=property_obj,
                unit=unit_obj,
                tenant=tenant_obj,
                code=code,
                amount=amount_val,
                date=date_obj,
                description=description,
            )
            allocate_payment_to_water_bills(payment)
            created_count += 1
        except ValueError as e:
            errors.append(f'Row {idx}: Invalid data format - {str(e)}')
        except Exception as e:
            errors.append(f'Row {idx}: {str(e)}')

    if validate_only:
        return JsonResponse({
            'success': len(errors) == 0,
            'valid_rows': valid_rows,
            'invalid_rows': len(errors),
            'errors': errors,
            'message': f'Validation complete: {valid_rows} valid, {len(errors)} invalid.',
        })

    message = f'Successfully created {created_count} water bill payments'
    if errors:
        message += f'. {len(errors)} errors: ' + '; '.join(errors[:3])

    return JsonResponse({
        'success': created_count > 0 or not errors,
        'count': created_count,
        'message': message,
        'errors': errors,
        'invalid_rows': len(errors),
    })
