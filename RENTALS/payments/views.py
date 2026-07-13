from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required

# Create your views here.
from django.shortcuts import render
from django.http import HttpResponse

from .models import Payment
from .forms import PaymentForm
from invoices.models import InvoicePayment
from properties.models import Property
from units.models import Unit
from leases.models import Lease
from invoices.services import allocate_payment_to_rent_invoices
import csv
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from PROPATIA.pagination import paginate_queryset


def resolve_active_lease_tenant(user, property_obj, unit_obj):
    if not unit_obj or unit_obj.user_id != user.id or unit_obj.property_id != property_obj.id:
        raise ValueError('Selected house/unit does not belong to the selected property')

    active_lease = Lease.objects.filter(
        user=user,
        unit=unit_obj,
        is_active=True,
    ).select_related('tenant').first()

    if not active_lease or not active_lease.tenant:
        raise ValueError(f'No active tenant found for house/unit "{unit_obj.name}" in property "{property_obj.name}"')

    return active_lease.tenant


@login_required
def payment_list(request):
    if request.method == "POST":
        form = PaymentForm(request.POST, user=request.user)
        if form.is_valid():
            payment = form.save(commit=False)
            try:
                payment.user = request.user
                payment.tenant = resolve_active_lease_tenant(request.user, payment.property, payment.unit)
                payment.save()
                allocate_payment_to_rent_invoices(payment)
                return redirect('payments:payment_list')
            except ValueError as e:
                form.add_error('unit', str(e))
    else:
        form = PaymentForm(user=request.user)

    # Get filters from GET parameters
    selected_property = request.GET.get('property', '')
    selected_status = request.GET.get('status', '')
    selected_start_date = request.GET.get('start_date', '')
    selected_end_date = request.GET.get('end_date', '')
    
    # Apply filters
    payments = Payment.objects.filter(user=request.user).select_related('property', 'unit', 'tenant').order_by('date', 'id')
    
    if selected_property:
        payments = payments.filter(property_id=selected_property)
    
    if selected_status:
        payments = payments.filter(status=selected_status)

    if selected_start_date:
        payments = payments.filter(date__gte=selected_start_date)

    if selected_end_date:
        payments = payments.filter(date__lte=selected_end_date)
    
    properties = Property.objects.filter(user=request.user)
    
    pagination = paginate_queryset(request, payments)
    page_payments = list(pagination['page_obj'])
    payment_ids = [payment.id for payment in page_payments]
    allocations_by_payment = {payment_id: [] for payment_id in payment_ids}

    if payment_ids:
        allocations = InvoicePayment.objects.filter(
            payment_id__in=payment_ids,
        ).select_related('invoice').order_by('invoice__due_date', 'invoice__id')
        for allocation in allocations:
            allocations_by_payment.setdefault(allocation.payment_id, []).append(allocation)

    for payment in page_payments:
        allocations = allocations_by_payment.get(payment.id, [])
        payment.allocated_amount = sum(allocation.amount_applied for allocation in allocations)
        payment.remaining_credit = payment.balance
        payment.allocation_summary = ', '.join(
            f'{allocation.invoice.due_date:%b %Y}: KES {allocation.amount_applied}'
            for allocation in allocations
        )
        payment.allocation_count = len(allocations)

    context = {
        'payments': pagination['page_obj'],
        'form': form,
        'properties': properties,
        'selected_property': selected_property,
        'selected_status': selected_status,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
    }
    context.update(pagination)
    return render(request, 'payments/payments_view.html', context)


@require_POST
def delete_payments(request):
    """Delete selected payments"""
    try:
        payment_ids = request.POST.getlist('payment_ids[]')
        
        if not payment_ids:
            return JsonResponse({'success': False, 'message': 'No payments selected'})
        
        deleted_count, _ = Payment.objects.filter(id__in=payment_ids).delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully deleted {deleted_count} payment(s)'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting payments: {str(e)}'
        })


@login_required
@require_POST
def upload_payments(request):
    """Upload payments from CSV or XLSX file"""
    def normalize_header(header):
        return ' '.join(
            str(header)
            .replace('\ufeff', '')
            .replace('*', '')
            .replace('\\', '')
            .lower()
            .replace('_', ' ')
            .split()
        )

    def get_row_value(row, *headers):
        for header in headers:
            value = row.get(normalize_header(header))
            if value is not None and str(value).strip() != '':
                return value
        return None

    def normalize_upload_row(headers, values):
        row = {}
        for idx, header in enumerate(headers):
            normalized_header = normalize_header(header)
            if not normalized_header:
                continue

            value = values[idx] if idx < len(values) else None
            current_value = row.get(normalized_header)
            if current_value is None or str(current_value).strip() == '':
                row[normalized_header] = value
        return row

    def parse_decimal(value):
        if value is None or str(value).strip() == '':
            raise ValueError('Missing amount')
        value_text = str(value).strip().replace(',', '')
        value_text = re.sub(r'[^\d.\-()]', '', value_text)
        if value_text.startswith('(') and value_text.endswith(')'):
            value_text = f'-{value_text[1:-1]}'
        if not value_text:
            raise ValueError('Invalid amount')
        try:
            return Decimal(value_text)
        except (InvalidOperation, ValueError):
            raise ValueError('Invalid amount')

    def parse_payment_date(value):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            value = value.strip()
            for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%m-%d-%Y', '%d.%m.%Y']:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        raise ValueError('Invalid date format')

    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file provided'})
    
    file = request.FILES['file']
    filename = file.name.lower()
    validate_only = request.POST.get('validate_only') == '1'
    
    try:
        rows = []
        
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            sample = content[:2048]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t;')
            except csv.Error:
                dialect = csv.excel_tab if '\t' in sample else csv.excel
            reader = csv.reader(content.splitlines(), dialect=dialect)
            headers = next(reader, [])
            rows = [
                normalize_upload_row(headers, row)
                for row in reader
                if any(value is not None and str(value).strip() != '' for value in row)
            ]
        elif filename.endswith('.xlsx'):
            try:
                import openpyxl
                from openpyxl import load_workbook
                import io
                
                wb = load_workbook(io.BytesIO(file.read()))
                ws = wb.active
                
                headers = [cell.value for cell in ws[1]]
                
                for row in ws.iter_rows(min_row=2, values_only=False):
                    row_dict = normalize_upload_row(headers, [cell.value for cell in row])
                    if any(row_dict.values()):
                        rows.append(row_dict)
            except ImportError:
                return JsonResponse({'success': False, 'message': 'openpyxl is not installed. Please use CSV format or contact admin.'})
        else:
            return JsonResponse({'success': False, 'message': 'Invalid file format. Please use CSV or XLSX'})
        
        if not rows:
            return JsonResponse({'success': False, 'message': 'File is empty'})
        
        created_count = 0
        errors = []
        valid_rows = 0
        
        for idx, row in enumerate(rows, start=2):
            try:
                row_lower = {normalize_header(k): v for k, v in row.items()}
                
                property_name = get_row_value(row_lower, 'property')
                house_number = get_row_value(
                    row_lower,
                    'HOUSE_NUMBER',
                    'house numner',
                    'house no',
                    'house',
                    'unit',
                    'unit number',
                )
                paid_in_date = get_row_value(row_lower, 'paid in date', 'date', 'payment date')
                code = get_row_value(row_lower, 'code', 'receipt', 'reference') or ''
                details = get_row_value(row_lower, 'details', 'description') or ''
                amount = get_row_value(row_lower, 'amount')
                
                if not property_name or not house_number or amount is None or not paid_in_date:
                    errors.append(f'Row {idx}: Missing required fields (PROPERTY, HOUSE_NUMBER, DATE, AMOUNT)')
                    continue
                
                property_obj = Property.objects.filter(name__iexact=property_name, user=request.user).first()
                if not property_obj:
                    errors.append(f'Row {idx}: Property "{property_name}" not found for this user')
                    continue
                
                house_number_str = str(house_number).strip()
                unit_obj = Unit.objects.filter(
                    property=property_obj,
                    name__iexact=house_number_str,
                    user=request.user,
                ).first()
                if not unit_obj:
                    errors.append(f'Row {idx}: House number "{house_number}" not found in property "{property_name}"')
                    continue

                try:
                    tenant_obj = resolve_active_lease_tenant(request.user, property_obj, unit_obj)
                except ValueError as e:
                    errors.append(f'Row {idx}: {str(e)}')
                    continue
                
                try:
                    amount_val = parse_decimal(amount)
                except ValueError:
                    errors.append(f'Row {idx}: Invalid amount format "{amount}"')
                    continue
                
                try:
                    date_obj = parse_payment_date(paid_in_date)
                except Exception as e:
                    errors.append(f'Row {idx}: Error parsing date - {str(e)}')
                    continue

                code = str(code).strip()
                description = str(details).strip()

                duplicate_exists = Payment.objects.filter(
                    user=request.user,
                    property=property_obj,
                    unit=unit_obj,
                    tenant=tenant_obj,
                    amount=amount_val,
                    date=date_obj,
                    code=code,
                    description=description,
                ).exists()
                if duplicate_exists:
                    errors.append(f'Row {idx}: Duplicate payment skipped for {property_obj.name} house {unit_obj.name} on {date_obj}')
                    continue

                valid_rows += 1

                if validate_only:
                    continue
                
                payment_data = {
                    'user': request.user,
                    'property': property_obj,
                    'unit': unit_obj,
                    'tenant': tenant_obj,
                    'code': code,
                    'amount': amount_val,
                    'date': date_obj,
                    'description': description,
                }
                
                payment = Payment.objects.create(**payment_data)
                allocate_payment_to_rent_invoices(payment)
                created_count += 1
                
            except ValueError as ve:
                errors.append(f'Row {idx}: Invalid data format - {str(ve)}')
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')
        
        if validate_only:
            return JsonResponse({
                'success': len(errors) == 0,
                'valid_rows': valid_rows,
                'invalid_rows': len(errors),
                'errors': errors,
                'message': f'Validation complete: {valid_rows} valid, {len(errors)} invalid.',
            })

        message = f'Successfully created {created_count} payments'
        if errors:
            message += f'. {len(errors)} errors: ' + '; '.join(errors[:3])

        response = {
            'success': created_count > 0 or not errors,
            'count': created_count,
            'message': message,
            'errors': errors,
            'invalid_rows': len(errors),
        }
        return JsonResponse(response)
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error processing file: {str(e)}'})
