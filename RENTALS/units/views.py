from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db.models import Prefetch, Q
import csv
import json
from decimal import Decimal, InvalidOperation
import io

# Create your views here.


from .models import Unit
from tenants.models import Tenant
from properties.models import Property
from django.contrib.auth.decorators import login_required
from PROPATIA.pagination import paginate_queryset
from .forms import UnitForm
from accounts.access_utils import get_accessible_properties, filter_units_by_accessible_properties

@login_required
def units_list(request):
    # 1. Handle POST (Form Submission)
    if request.method == "POST":
        form = UnitForm(request.POST, user=request.user)
        if form.is_valid():
            unit = form.save(commit=False)
            unit.user = request.user
            unit.save()
            return redirect('units:units_list')
    else:
        form = UnitForm(user=request.user)

    # Filter units to only those in accessible properties
    accessible_props = get_accessible_properties(request.user)
    units = Unit.objects.filter(
        property__in=accessible_props
    ).select_related('property').prefetch_related(
        Prefetch('tenants', queryset=Tenant.objects.order_by('last_name'), to_attr='prefetched_tenants')
    ).order_by('property__name', 'name')
    
    property_filter = request.GET.get('property')
    status_filter = request.GET.get('status')
    
    if property_filter:
        units = units.filter(property_id=property_filter)
    if status_filter:
        units = units.filter(status=status_filter)
            
    # Populate tenant_name from prefetched tenants (via active lease)
    from leases.models import Lease
    unit_ids = [u.id for u in units]
    active_leases = Lease.objects.filter(unit_id__in=unit_ids, is_active=True).select_related('tenant')
    lease_by_unit = {l.unit_id: l for l in active_leases}
    for unit in units:
        lease = lease_by_unit.get(unit.id)
        if lease and lease.tenant:
            unit.tenant_name = f"{lease.tenant.first_name} {lease.tenant.last_name}"
        elif unit.prefetched_tenants:
            t = unit.prefetched_tenants[0]
            unit.tenant_name = f"{t.first_name} {t.last_name}"
        else:
            unit.tenant_name = None

    today_date = timezone.now().date()
    pagination = paginate_queryset(request, units)

    context = {
        'units': pagination['page_obj'],
        'form': form,
        'properties': accessible_props,
        'tenants': Tenant.objects.filter(unit__isnull=True).order_by('first_name', 'last_name'),
        'today_date': today_date,
        'selected_property': property_filter,
        'selected_status': status_filter,
    }
    context.update(pagination)
    return render(request, 'units/units_view.html', context)


def assign_tenant(request, pk):
    """Assign a tenant to a unit"""
    accessible_props = get_accessible_properties(request.user)
    unit = get_object_or_404(Unit.objects.filter(property__in=accessible_props), pk=pk)

    if request.method == 'POST':
        tenant_id = request.POST.get('tenant_id')
        start_date = request.POST.get('start_date')
        deposit_held = request.POST.get('deposit_held', 0)
        if tenant_id:
            try:
                from leases.models import Lease
                tenant = Tenant.objects.filter(id=tenant_id).exclude(leases__is_active=True).first()
                if not tenant:
                    return JsonResponse({'success': False, 'message': 'Selected tenant is not available or does not belong to you.'}, status=400)
                if tenant.leases.filter(is_active=True).exists():
                    return JsonResponse({'success': False, 'message': 'Tenant already has an active lease.'}, status=400)
                if Lease.objects.filter(unit=unit, is_active=True).exists():
                    return JsonResponse({'success': False, 'message': f'Unit {unit.name} already has an active lease.'}, status=400)
                lease = Lease.objects.create(
                    user=request.user,
                    tenant=tenant,
                    unit=unit,
                    start_date=start_date or timezone.now().date(),
                    monthly_rent=unit.rent_amount,
                    deposit_held=deposit_held,
                    is_active=True
                )
                tenant.unit = unit
                tenant.save()
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': f'Lease created: {tenant.first_name} {tenant.last_name} assigned to {unit.name}'})
                else:
                    return redirect('units:units_list')
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)}, status=400)
        else:
            return JsonResponse({'success': False, 'message': 'Please select a tenant'}, status=400)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        tenants = Tenant.objects.filter(unit__isnull=True).exclude(leases__is_active=True).order_by('last_name')
        return JsonResponse({
            'unit_id': unit.id,
            'unit_name': unit.name,
            'tenants': list(tenants.values('id', 'first_name', 'last_name')),
            'message': f'{tenants.count()} tenant(s) available for {unit.name}'
        })


def detach_tenant(request, pk):
    """Detach tenant from a unit"""
    accessible_props = get_accessible_properties(request.user)
    unit = get_object_or_404(Unit.objects.filter(property__in=accessible_props), pk=pk)
    
    if request.method == 'POST':
        from leases.models import Lease
        lease = Lease.objects.filter(unit=unit, is_active=True).first()
        if lease:
            lease.is_active = False
            lease.save()
            if hasattr(lease.tenant, 'unit'):
                lease.tenant.unit = None
                lease.tenant.save()
        
        return JsonResponse({'success': True, 'message': f'Tenant detached from {unit.name}'})


@require_POST
def delete_units(request):
    """Delete selected units"""
    unit_ids = request.POST.getlist('unit_ids[]')
    
    if unit_ids:
        accessible_props = get_accessible_properties(request.user)
        Unit.objects.filter(id__in=unit_ids, property__in=accessible_props).delete()
        return JsonResponse({'success': True, 'message': f'{len(unit_ids)} unit/s deleted successfully'})
    
    return JsonResponse({'success': False, 'message': 'No units selected'})


@require_POST
def upload_units(request):
    """Upload units from CSV or XLSX file"""
    def normalize_header(header):
        return ' '.join(str(header).lower().replace('_', ' ').split()).replace(' ', '_')

    def normalize_text(value):
        if value is None:
            return ''
        return str(value).strip()

    def parse_decimal(value):
        if value is None or str(value).strip() == '':
            raise ValueError('Missing rent_amount')
        try:
            return Decimal(str(value).replace(',', '').strip())
        except (InvalidOperation, ValueError):
            raise ValueError('Invalid rent_amount')

    def read_csv_rows(uploaded_file):
        content = uploaded_file.read().decode('utf-8-sig')
        sample = content[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',\t;')
        except csv.Error:
            dialect = csv.excel_tab if '\t' in sample else csv.excel

        reader = csv.DictReader(io.StringIO(content), dialect=dialect)
        normalized_fieldnames = [normalize_header(header) for header in (reader.fieldnames or [])]
        reader.fieldnames = normalized_fieldnames
        return [repair_unit_row(row) for row in reader]

    def read_xlsx_rows(uploaded_file):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError('openpyxl is not installed. Please use CSV format or contact admin.')

        wb = load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)
        ws = wb.active
        headers = [normalize_header(cell.value) for cell in ws[1]]
        rows = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for idx, header in enumerate(headers):
                if header:
                    row_dict[header] = row[idx] if idx < len(row) else None
            if any(value is not None and str(value).strip() != '' for value in row_dict.values()):
                rows.append(row_dict)
        return rows

    def repair_unit_row(row):
        extra_values = row.get(None, [])
        row = {normalize_header(k): v for k, v in row.items() if k is not None}

        if extra_values and row.get('rent_amount') and row.get('description'):
            rent_part = normalize_text(row.get('rent_amount'))
            decimal_part = normalize_text(row.get('description'))
            if rent_part.isdigit() and decimal_part.isdigit():
                row['rent_amount'] = f'{rent_part},{decimal_part}'
                row['description'] = row.get('status', '')
                row['status'] = extra_values[0] if extra_values else ''

        return row

    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file provided'})
    
    file = request.FILES['file']
    filename = file.name.lower()
    
    try:
        rows = []
        
        if filename.endswith('.csv'):
            rows = read_csv_rows(file)
        elif filename.endswith('.xlsx'):
            try:
                rows = read_xlsx_rows(file)
            except RuntimeError as e:
                return JsonResponse({'success': False, 'message': str(e)})
        else:
            return JsonResponse({'success': False, 'message': 'Invalid file format. Please use CSV or XLSX'})
        
        if not rows:
            return JsonResponse({'success': False, 'message': 'File is empty'})
        
        accessible_props = get_accessible_properties(request.user)
        created_count = 0
        errors = []
        
        for idx, row in enumerate(rows, start=2):
            try:
                row_lower = {normalize_header(k): v for k, v in row.items()}
                
                property_name = normalize_text(row_lower.get('property'))
                unit_name = normalize_text(row_lower.get('name'))
                rent_amount = row_lower.get('rent_amount')
                
                if not property_name or not unit_name or rent_amount is None or str(rent_amount).strip() == '':
                    errors.append(f'Row {idx}: Missing required fields (property, name, rent_amount)')
                    continue
                
                property_obj = accessible_props.filter(name__iexact=property_name).first()
                if not property_obj:
                    errors.append(f'Row {idx}: Property "{property_name}" not found or not accessible')
                    continue
                
                try:
                    rent_amount_val = parse_decimal(rent_amount)
                except ValueError as ve:
                    errors.append(f'Row {idx}: Invalid rent_amount "{rent_amount}"')
                    continue

                status = normalize_text(row_lower.get('status')).lower() or 'vacant'
                if status not in dict(Unit.STATUS_CHOICES):
                    errors.append(f'Row {idx}: Invalid status "{row_lower.get("status")}". Use occupied or vacant')
                    continue
                
                unit_data = {
                    'user': request.user,
                    'property': property_obj,
                    'name': unit_name,
                    'rent_amount': rent_amount_val,
                    'description': normalize_text(row_lower.get('description')),
                    'status': status,
                }
                
                Unit.objects.create(**unit_data)
                created_count += 1
                
            except ValueError as ve:
                errors.append(f'Row {idx}: Invalid data format - {str(ve)}')
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')
        
        message = f'Successfully created {created_count} units'
        if errors:
            message += f'. {len(errors)} errors: ' + '; '.join(errors[:3])

        response = {
            'success': created_count > 0 or not errors,
            'count': created_count,
            'message': message,
            'errors': errors,
            'invalid_rows': len(errors),
        }
        return JsonResponse(response)
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error processing file: {str(e)}'})
